from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.plotarea import DataTable

sheet_name = "mySheet"

wb = Workbook()
ws = wb.active
ws.title = sheet_name

data = [["Training domain \ Methods", "FFT-SVM", "FFT-MLP", "FFT-DNN", "WDCNN", "TICNN", "Ensemble TICNN"],
        ["1→2", 0.686, 0.821, 0.822 ,0.992, 0.991, 0.995],
        ["1→3", 0.600, 0.856, 0.826, 0.910, 0.907, 0.911],
        ["2→1", 0.732, 0.715, 0.723, 0.951, 0.974, 0.976],
        ["2→3", 0.676, 0.824, 0.770, 0.915, 0.988, 0.994],
        ["3→1", 0.684, 0.818, 0.769, 0.781, 0.892, 0.902],
        ["3→2", 0.620, 0.790, 0.773, 0.851, 0.976, 0.987],
        ["AVG", 0.666, 0.804, 0.781, 0.900, 0.955, 0.961]]

row_num = len(data)
col_num = len(data[0])
#%% 
# add data to worksheet line by line
for row in data:
    ws.append(row)

# change display style
for x in range(2,row_num+1):
    for y in range(2,col_num+1):
        ws.cell(row=x, column=y).number_format = "0.0%"

#%% draw bar chart

barChart = BarChart()
barChart.type = "col"
barChart.style = 10
barChart.title = "version " + sheet_name
barChart.x_axis.title = "Loading domain"
barChart.y_axis.title = "Accuracy (%)"
barChart.varColors = True

# show bar chart data table
barChart.plot_area.dTable = DataTable()
barChart.plot_area.dTable.showHorzBorder = True
barChart.plot_area.dTable.showVertBorder = True
barChart.plot_area.dTable.showOutline = True
barChart.plot_area.dTable.showKeys = True

# chart size
barChart.width = 20
barChart.height = 12

# axis scale
barChart.y_axis.scaling.min = 0.5
barChart.y_axis.scaling.max = 1
data = Reference(ws, min_row=1, min_col=2, max_row=row_num, max_col=col_num)
cats = Reference(ws, min_row=2, min_col=1, max_row=row_num)
barChart.add_data(data, titles_from_data=True, from_rows=False)
barChart.set_categories(cats)
anchor = ws.cell(row=row_num+2, column=1).coordinate
ws.add_chart(barChart, anchor=anchor)   # add_chart(chart, "cell")

wb.save(sheet_name + ".xlsx")